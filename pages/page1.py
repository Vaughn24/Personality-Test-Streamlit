import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import os

from streamlit_extras.switch_page_button import switch_page
def remove_sidebar():

    st.set_page_config(initial_sidebar_state="collapsed")
    st.markdown(
        """
    <style>
        [data-testid="collapsedControl"] {
            display: none
        }
    </style>
    """,
        unsafe_allow_html=True,
    )

if __name__ == '__main__':
    remove_sidebar()
    st.title('Data Privacy')
    st.write("By filling-out and submitting this form, you hereby authorize to use, collect, and process the information for legitimate purposes specifically for your participation on this survey, and allow authorized personnel to process the information pursuant to the Data Privacy policies.")
    answer = st.radio("", ("Yes, I hereby allow/authorize to use, collect, and process the information for legitimate purposes specifically for the academic study and allow authorized personnel to process the information.", "No"))
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    style = """<style>
    .row-widget.stButton {
    margin-top: 2rem}
    </style>"""
    st.markdown(style, unsafe_allow_html=True)


    with col6:

        if st.button("Next Page"):
            if answer == "No":
                switch_page("page5")
            else:
                # Save the answer to an Excel file
                columns = ['Would you like to proceed?', 'First Name', 'Last Name', 'Birth Date', 'Gender',
                           'Email Address']
                df = pd.DataFrame(columns=columns)
                default = [answer, ' ', ' ', ' ', ' ', ' ']
                df.loc[len(df)] = default
                excel_file_name = 'user_answers.xlsx'
                wb = openpyxl.Workbook()
                if os.path.exists(excel_file_name):
                    os.remove(excel_file_name)
                    wb.save(excel_file_name)
                else:
                    wb.save(excel_file_name)
                # Check if the file already exists, and if so, delete it
                with pd.ExcelWriter(excel_file_name, mode='a', engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False)

                # Remove the first sheet
                wb = openpyxl.load_workbook(excel_file_name)
                sheet = wb.sheetnames
                pfd = wb['Sheet']
                wb.remove(pfd)
                wb.save(excel_file_name)

                switch_page("page2")


    with col1:
        if st.button("Prev Page"):
            switch_page("app")






